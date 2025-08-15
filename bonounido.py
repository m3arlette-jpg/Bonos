# 📦 Importamos las bibliotecas necesarias
import streamlit as st
import pandas as pd
import re
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile

# 🎨 Configuración inicial
st.set_page_config(
    page_title="VALIDACION DE CARTAS VEAB💲",
    page_icon="https://i.pinimg.com/474x/85/75/02/85750251513131f033ccadb1ee162581.jpg",
    layout="wide"
)
st.markdown(
    """
    <style>
        .stApp {
            background-color: #263576;
        }
    </style>
    """,
    unsafe_allow_html=True
)


st.title("💲 VALIDACION DE CARTAS VEAB PDF vs CSV")
st.markdown(
    """
    <div style='text-align: center'>
        <img src='https://i.pinimg.com/originals/8f/b7/bd/8fb7bdb708916be41b4789bdad9c27cb.gif' width='150'/>
    </div>
    """,
    unsafe_allow_html=True
)



# 🔐 Protección con contraseña
def verificar_acceso():
    st.sidebar.header("🔒 Acceso restringido")
    password = st.sidebar.text_input("Ingresa la contraseña", type="password")
    return password == "miclave123"  # Cambia esto por tu clave

if not verificar_acceso():
    st.warning("🔐 Esta aplicación está protegida. Ingresa la contraseña en la barra lateral.")
    st.stop()


# 🗂️ Pestañas principales
tab_acciones_es, tab_acciones_en, tab_bono_es, tab_bono_en = st.tabs([
    "🇪🇸 Acciones", "🇺🇸 Virtual Shares", "🇪🇸 Bono Diferido", "🇺🇸 Deferred Bonus"
])

# ─────────────────────────────────────────────────────────────
# 🇪🇸 Comparador de Acciones Español
with tab_acciones_es:
    st.header("📂 Acciones")

    def limpiar_es(valor):
        return str(valor).replace(",", "").replace("\xa0", "").replace("\u200b", "").replace(" ", "").replace("%", "").strip()

    def comparar_valores_es(pdf_valor, csv_valor):
        pdf_valor = limpiar_es(pdf_valor)
        csv_valor = limpiar_es(csv_valor)
        try:
            pdf_float = round(float(pdf_valor), 2)
            csv_float = round(float(csv_valor), 2)
            return abs(pdf_float - csv_float) < 0.01
        except ValueError:
            return pdf_valor == csv_valor

    def extraer_datos_acciones_es(texto):
        acciones = re.search(r'asignado.*?([\d,]+)', texto, re.IGNORECASE)
        factor = re.search(r'reportas.*?:\s*([\d]+)', texto, re.IGNORECASE)
        porcentaje = re.search(r'corresponden.*?:\s*([\d]+)', texto, re.IGNORECASE)
        salario = re.search(r'2024.*?:\s*([\d,]+(?:\.\d{2})?)', texto, re.IGNORECASE)
        equivalente = re.search(r'equivalente a\s*([\d,]+(?:\.\d{2})?)', texto, re.IGNORECASE)
        if acciones and factor and porcentaje and salario and equivalente:
            return (
                limpiar_es(acciones.group(1)),
                limpiar_es(factor.group(1)),
                limpiar_es(porcentaje.group(1)),
                "{:.2f}".format(float(limpiar_es(salario.group(1)))),
                "{:.2f}".format(float(limpiar_es(equivalente.group(1))))
            )
        return None

    def extraer_nombre_acciones_es(texto):
        lineas = texto.splitlines()
        for i, linea in enumerate(lineas):
            if re.search(r'^Junio\s+\d{4}$', linea.strip(), re.IGNORECASE):
                for j in range(i + 1, len(lineas)):
                    siguiente = lineas[j].strip()
                    if siguiente:
                        return siguiente
        return None

    def procesar_acciones_es(df, pdf_files, columnas):
        df[columnas[1:]] = df[columnas[1:]].applymap(limpiar_es)
        df[columnas[0]] = df[columnas[0]].astype(str).str.upper().str.strip()
        errores_por_fila = {}
        comentarios = {}
        iconos_df = df.copy()
        iconos_df["Origen PDF"] = ""
        notas = []
        procesados = []

        for file in pdf_files:
            reader = PdfReader(file)
            texto = ''.join(page.extract_text() for page in reader.pages if page.extract_text())
            if not texto.strip():
                continue
            nombre_pdf = extraer_nombre_acciones_es(texto)
            if not nombre_pdf:
                continue
            nombre_pdf = nombre_pdf.upper().strip()
            if nombre_pdf not in df[columnas[0]].values:
                continue
            idx = df[df[columnas[0]] == nombre_pdf].index[0]
            fila = df.loc[idx]
            datos = extraer_datos_acciones_es(texto)
            if not datos:
                continue
            errores = []
            for campo, extraido, esperado in zip(columnas[1:], datos, [str(fila[col]) for col in columnas[1:]]):
                if not comparar_valores_es(extraido, esperado):
                    errores.append(campo)
                    comentarios[(idx, campo)] = f"{campo}: En el EXCEL: {esperado}// En el PDF: {extraido}"
                    iconos_df.at[idx, campo] = f"❌ {fila[campo]}"
                else:
                    iconos_df.at[idx, campo] = f"✅ {fila[campo]}"
            if errores:
                errores_por_fila[idx] = errores
            iconos_df.at[idx, "Origen PDF"] = file.name
            procesados.append(idx)

        for idx in iconos_df.index:
            fila_notas = [comentarios[(idx, col)] for col in iconos_df.columns if (idx, col) in comentarios]
            notas.append(" | ".join(fila_notas))
        iconos_df["Notas"] = notas

        def resaltar(row):
            idx = row.name
            return ['background-color: #FFCCCC' if col in errores_por_fila.get(idx, []) else ''
                    for col in iconos_df.columns]

        iconos_filtrados = iconos_df.loc[procesados] if procesados else pd.DataFrame()
        st.subheader("📊 Resultados comparados")
        if not iconos_filtrados.empty:
            st.dataframe(iconos_filtrados.style.apply(resaltar, axis=1), use_container_width=True)
        else:
            st.warning("⚠️ No se encontraron coincidencias válidas entre los PDFs y los nombres del CSV.")

        # 💾 Botón para exportar Excel con errores marcados
        if st.button("📥 Generar Excel con errores resaltados", key="descargar_acciones_es"):
            if procesados:
                export_df = df.loc[procesados].copy()
                export_df["Origen PDF"] = iconos_df.loc[procesados]["Origen PDF"]
                export_df["Notas"] = iconos_df.loc[procesados]["Notas"]
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    ruta_excel = tmp.name
                    export_df.to_excel(ruta_excel, index=False)
                    wb = load_workbook(ruta_excel)
                    ws = wb.active
                    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    for idx in errores_por_fila:
                        excel_row = list(procesados).index(idx) + 2  # +2 por encabezado
                        for col in errores_por_fila[idx]:
                            col_idx = list(export_df.columns).index(col) + 1
                            ws.cell(row=excel_row, column=col_idx).fill = red_fill
                    wb.save(ruta_excel)
                with open(ruta_excel, "rb") as f:
                    st.download_button("📥 Descargar comaparacion_accionesESP.xlsx", f, file_name="comaparacion_accionesESP.xlsx")
            else:
                st.error("⚠️ No hay filas válidas para exportar.")
    

    columnas_acciones_es = ['Nombre', 'Acciones', 'Factor financiero', 'Target', 'Salario Diario', 'Acciones MXN']
    csv_file_es = st.file_uploader("📂 Sube tu archivo CSV", type=["csv"], key="csv_acciones_es")
    pdf_files_es = st.file_uploader("📥 Sube tus archivos PDF", type=["pdf"], accept_multiple_files=True, key="pdf_acciones_es")
    if csv_file_es and pdf_files_es:
        df_es = pd.read_csv(csv_file_es)
        if all(col in df_es.columns for col in columnas_acciones_es):
            procesar_acciones_es(df_es, pdf_files_es, columnas_acciones_es)
        else:
            st.error(f"⚠️ El CSV debe tener las columnas: {columnas_acciones_es}")

# ─────────────────────────────────────────────────────────────
# 🇺🇸 Comparador de Acciones Inglés
with tab_acciones_en:
    st.header("📂 Virtual Shares")

    def limpiar_en(valor):
        return str(valor).replace(",", "").replace("\xa0", "").replace("\u200b", "").replace(" ", "").replace("%", "").strip()

    def comparar_valores_en(pdf_valor, csv_valor):
        pdf_valor = limpiar_en(pdf_valor)
        csv_valor = limpiar_en(csv_valor)
        try:
            pdf_float = round(float(pdf_valor), 2)
            csv_float = round(float(csv_valor), 2)
            return abs(pdf_float - csv_float) < 0.01
        except ValueError:
            return pdf_valor == csv_valor

    def extraer_datos_acciones_en(texto):
        acciones = re.search(r'assigned\s+([\d,\.]+)', texto, re.IGNORECASE)
        factor = re.search(r'financial factor.*?(\d+)', texto, re.IGNORECASE)
        porcentaje = re.search(r'target bonus.*?(\d+(?:\.\d+)?)', texto, re.IGNORECASE)
        salario = re.search(r'December \d{4}.*?([\d,]+\.\d{2})', texto, re.IGNORECASE)
        equivalente = re.search(r'equivalent to\s+([\d,\.]+)', texto, re.IGNORECASE)
        if acciones and factor and porcentaje and salario and equivalente:
            return (
                limpiar_en(acciones.group(1)),
                limpiar_en(factor.group(1)),
                limpiar_en(porcentaje.group(1)),
                "{:.2f}".format(float(limpiar_en(salario.group(1)))),
                "{:.2f}".format(float(limpiar_en(equivalente.group(1))))
            )
        return None

    def extraer_nombre_acciones_en(texto):
        lineas = texto.splitlines()
        for i, linea in enumerate(lineas):
            normalizada = re.sub(r'\s+', '', linea.strip())
            if re.search(r'^May,\d{4}$', normalizada, re.IGNORECASE):
                for j in range(i + 1, len(lineas)):
                    siguiente = lineas[j].strip()
                    if siguiente:
                        return siguiente
        return None

    def procesar_acciones_en(df, pdf_files, columnas):
        df[columnas[1:]] = df[columnas[1:]].applymap(limpiar_en)
        df[columnas[0]] = df[columnas[0]].astype(str).str.upper().str.strip()
        errores_por_fila = {}
        comentarios = {}
        iconos_df = df.copy()
        iconos_df["PDF SOURCE"] = ""
        notas = []
        procesados = []

        for file in pdf_files:
            reader = PdfReader(file)
            texto = ''.join(page.extract_text() for page in reader.pages if page.extract_text())
            if not texto.strip():
                continue
            nombre_pdf = extraer_nombre_acciones_en(texto)
            if not nombre_pdf:
                continue
            nombre_pdf = nombre_pdf.upper().strip()
            if nombre_pdf not in df[columnas[0]].values:
                continue
            idx = df[df[columnas[0]] == nombre_pdf].index[0]
            fila = df.loc[idx]
            datos = extraer_datos_acciones_en(texto)
            if not datos:
                continue
            errores = []
            for campo, extraido, esperado in zip(columnas[1:], datos, [str(fila[col]) for col in columnas[1:]]):
                if not comparar_valores_en(extraido, esperado):
                    errores.append(campo)
                    comentarios[(idx, campo)] = f"{campo}: In CSV: {esperado}// In PDF: {extraido}"
                    iconos_df.at[idx, campo] = f"❌ {fila[campo]}"
                else:
                    iconos_df.at[idx, campo] = f"✅ {fila[campo]}"
            if errores:
                errores_por_fila[idx] = errores
            iconos_df.at[idx, "PDF SOURCE"] = file.name
            procesados.append(idx)

        for idx in iconos_df.index:
            fila_notas = [comentarios[(idx, col)] for col in iconos_df.columns if (idx, col) in comentarios]
            notas.append(" | ".join(fila_notas))
        iconos_df["NOTES"] = notas

        def resaltar(row):
            idx = row.name
            return ['background-color: #FFCCCC' if col in errores_por_fila.get(idx, []) else ''
                    for col in iconos_df.columns]

        iconos_filtrados = iconos_df.loc[procesados] if procesados else pd.DataFrame()
        st.subheader("📊 Comparison Results")
        if not iconos_filtrados.empty:
            st.dataframe(iconos_filtrados.style.apply(resaltar, axis=1), use_container_width=True)
        else:
            st.warning("⚠️ No valid matches found between PDFs and CSV names.")


        # 💾 Botón para exportar Excel con errores marcados
        if st.button("📥 Create Excel", key="descargar_acciones_ING"):
            if procesados:
                export_df = df.loc[procesados].copy()
                export_df["PDF SOURCE"] = iconos_df.loc[procesados]["PDF SOURCE"]
                export_df["NOTES"] = iconos_df.loc[procesados]["NOTES"]
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    ruta_excel = tmp.name
                    export_df.to_excel(ruta_excel, index=False)
                    wb = load_workbook(ruta_excel)
                    ws = wb.active
                    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    for idx in errores_por_fila:
                        excel_row = list(procesados).index(idx) + 2  # +2 por encabezado
                        for col in errores_por_fila[idx]:
                            col_idx = list(export_df.columns).index(col) + 1
                            ws.cell(row=excel_row, column=col_idx).fill = red_fill
                    wb.save(ruta_excel)
                with open(ruta_excel, "rb") as f:
                    st.download_button("📥 Download Compare_VirtualShares.xlsx", f, file_name="Compare_VirtualShares.xlsx")
            else:
                st.error("⚠️ No hay filas válidas para exportar.")    

    columnas_acciones_en = ['NAME', 'VIRTUAL SHARES', 'FINANCIAL FACTOR', 'TARGET BONUS', 'ANNUAL SALARY', 'VIRTUAL SHARES MXN']
    csv_file_en = st.file_uploader("📂 Upload your CSV file", type=["csv"], key="csv_acciones_en")
    pdf_files_en = st.file_uploader("📥 Upload your PDF files", type=["pdf"], accept_multiple_files=True, key="pdf_acciones_en")
    if csv_file_en and pdf_files_en:
        df_en = pd.read_csv(csv_file_en)
        if all(col in df_en.columns for col in columnas_acciones_en):
            procesar_acciones_en(df_en, pdf_files_en, columnas_acciones_en)
        else:
            st.error(f"⚠️ Your CSV must contain the following columns: {columnas_acciones_en}")

# ─────────────────────────────────────────────────────────────
# 🇪🇸 Comparador Bono Diferido Español
with tab_bono_es:
    st.header("📂 Bono Diferido")

    def limpiar_bono_es(valor):
        return str(valor).replace(",", "").replace("\xa0", "").replace("\u200b", "").replace(" ", "").replace("%", "").strip()

    def comparar_valores_bono_es(pdf_valor, csv_valor):
        pdf_valor = limpiar_bono_es(pdf_valor)
        csv_valor = limpiar_bono_es(csv_valor)
        try:
            return round(float(pdf_valor), 2) == round(float(csv_valor), 2)
        except ValueError:
            return pdf_valor == csv_valor

    def extraer_datos_bono_es(texto):
        bono = re.search(r'asignado.*?([\d,.]+)', texto, re.IGNORECASE)
        factor = re.search(r'reportas.*?:\s*([\d]+)', texto, re.IGNORECASE)
        porcentaje = re.search(r'corresponden.*?:\s*([\d]+)', texto, re.IGNORECASE)
        salario = re.search(r'2024.*?:\s*([\d,]+(?:\.\d{2})?)', texto, re.IGNORECASE)
        if bono and factor and porcentaje and salario:
            return limpiar_bono_es(bono.group(1)), limpiar_bono_es(factor.group(1)), limpiar_bono_es(porcentaje.group(1)), "{:.2f}".format(float(limpiar_bono_es(salario.group(1))))
        return None

    def extraer_nombre_bono_es(texto):
        lineas = texto.splitlines()
        for i, linea in enumerate(lineas):
            if re.search(r'^Mayo\s+\d{4}$', linea.strip(), re.IGNORECASE):
                for j in range(i + 1, len(lineas)):
                    siguiente = lineas[j].strip()
                    if siguiente:
                        return siguiente
        return None

    def comparar_bono_es(csv_file, pdf_files, columnas):
        df = pd.read_csv(csv_file)
        if not all(col in df.columns for col in columnas):
            st.error(f"⚠️ El CSV debe tener las columnas: {columnas}")
            return

        df[columnas[1:]] = df[columnas[1:]].applymap(limpiar_bono_es)
        df[columnas[0]] = df[columnas[0]].astype(str).str.upper().str.strip()

        errores_por_fila = {}
        comentarios = {}
        iconos_df = df.copy()
        iconos_df["ORIGEN PDF"] = ""
        notas = []
        procesados = []

        for file in pdf_files:
            reader = PdfReader(file)
            texto = ''.join(page.extract_text() for page in reader.pages if page.extract_text())
            if not texto.strip():
                continue

            nombre_pdf = extraer_nombre_bono_es(texto)
            if not nombre_pdf:
                continue

            nombre_pdf = nombre_pdf.upper().strip()
            if nombre_pdf not in df[columnas[0]].values:
                continue

            idx = df[df[columnas[0]] == nombre_pdf].index[0]
            fila = df.loc[idx]
            datos = extraer_datos_bono_es(texto)
            if not datos:
                continue

            errores = []
            for campo, extraido, esperado in zip(columnas[1:], datos, [fila[col] for col in columnas[1:]]):
                if not comparar_valores_bono_es(extraido, esperado):
                    errores.append(campo)
                    comentarios[(idx, campo)] = f"{campo}: En el EXCEL: {esperado}// En el PDF: {extraido}"
                    iconos_df.at[idx, campo] = f"❌ {fila[campo]}"
                else:
                    iconos_df.at[idx, campo] = f"✅ {fila[campo]}"

            if errores:
                errores_por_fila[idx] = errores
            else:
                for campo in columnas[1:]:
                    iconos_df.at[idx, campo] = f"✅ {fila[campo]}"

            iconos_df.at[idx, "ORIGEN PDF"] = file.name
            procesados.append(idx)

        for idx in iconos_df.index:
            fila_notas = [comentarios[(idx, col)] for col in iconos_df.columns if (idx, col) in comentarios]
            notas.append(" | ".join(fila_notas))
        iconos_df["NOTAS"] = notas

        def resaltar(row):
            idx = row.name
            return ['background-color: #FFCCCC' if col in errores_por_fila.get(idx, []) else ''
                    for col in iconos_df.columns]

        iconos_filtrados = iconos_df.loc[procesados] if procesados else pd.DataFrame()
        st.subheader("📊 Resultados comparados")

        if not iconos_filtrados.empty:
            st.dataframe(iconos_filtrados.style.apply(resaltar, axis=1), use_container_width=True)
        else:
            st.warning("⚠️ No se encontraron coincidencias válidas entre los PDFs y los nombres del CSV.")


        # 💾 Botón para exportar Excel con errores marcados
        if st.button("📥 Generar Excel con errores resaltados", key="descargar_bono_es"):
            if procesados:
                export_df = df.loc[procesados].copy()
                export_df["ORIGEN PDF"] = iconos_df.loc[procesados]["ORIGEN PDF"]
                export_df["NOTAS"] = iconos_df.loc[procesados]["NOTAS"]
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    ruta_excel = tmp.name
                    export_df.to_excel(ruta_excel, index=False)
                    wb = load_workbook(ruta_excel)
                    ws = wb.active
                    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    for idx in errores_por_fila:
                        excel_row = list(procesados).index(idx) + 2  # +2 por encabezado
                        for col in errores_por_fila[idx]:
                            col_idx = list(export_df.columns).index(col) + 1
                            ws.cell(row=excel_row, column=col_idx).fill = red_fill
                    wb.save(ruta_excel)
                with open(ruta_excel, "rb") as f:
                    st.download_button("📥 Descargar comaparacion_bonoESP.xlsx", f, file_name="comaparacion_bonoESP.xlsx")
            else:
                st.error("⚠️ No hay filas válidas para exportar.")
        

    columnas_bono_es = ['NOMBRE', 'BONO DIFERIDO', 'FACTOR FINANCIERO', 'DIAS BONO', 'SALARIO DIARIO']
    csv_file_bono_es = st.file_uploader("📂 Sube tu archivo CSV", type=["csv"], key="csv_bono_es")
    pdf_files_bono_es = st.file_uploader("📥 Sube tus PDFs", type=["pdf"], accept_multiple_files=True, key="pdf_bono_es")
    if csv_file_bono_es and pdf_files_bono_es:
        comparar_bono_es(csv_file_bono_es, pdf_files_bono_es, columnas_bono_es)

# ─────────────────────────────────────────────────────────────
# 🇺🇸 Comparador Bono Diferido Inglés
with tab_bono_en:
    st.header("📂 Deferred Bonus")

    def limpiar_bono_en(valor):
        return str(valor).replace(",", "").replace("\xa0", "").replace("\u200b", "").replace(" ", "").replace("%", "").strip()

    def comparar_valores_bono_en(pdf_valor, csv_valor):
        pdf_valor = limpiar_bono_en(pdf_valor)
        csv_valor = limpiar_bono_en(csv_valor)
        try:
            return round(float(pdf_valor), 2) == round(float(csv_valor), 2)
        except ValueError:
            return pdf_valor == csv_valor

    def extraer_datos_bono_en(texto):
        bono = re.search(r'assigned\s+([\d,\.]+)', texto, re.IGNORECASE)
        factor = re.search(r'financial factor.*?(\d+)', texto, re.IGNORECASE)
        porcentaje = re.search(r'target bonus.*?(\d+(?:\.\d+)?)', texto, re.IGNORECASE)
        salario = re.search(r'December \d{4}.*?([\d,]+\.\d{2})', texto, re.IGNORECASE)
        if bono and factor and porcentaje and salario:
            return limpiar_bono_en(bono.group(1)), limpiar_bono_en(factor.group(1)), limpiar_bono_en(porcentaje.group(1)), "{:.2f}".format(float(limpiar_bono_en(salario.group(1))))
        return None

    def extraer_nombre_bono_en(texto):
        lineas = texto.splitlines()
        for i, linea in enumerate(lineas):
            normalizada = re.sub(r'\s+', '', linea.strip())
            if re.search(r'^May,\d{4}$', normalizada, re.IGNORECASE):
                for j in range(i + 1, len(lineas)):
                    siguiente = lineas[j].strip()
                    if siguiente:
                        return siguiente
        return None

    def comparar_bono_en(csv_file, pdf_files, columnas):
        df = pd.read_csv(csv_file)
        if not all(col in df.columns for col in columnas):
            st.error(f"⚠️ Your CSV must contain the following columns: {columnas}")
            return

        df[columnas[1:]] = df[columnas[1:]].applymap(limpiar_bono_en)
        df[columnas[0]] = df[columnas[0]].astype(str).str.upper().str.strip()

        errores_por_fila = {}
        comentarios = {}
        iconos_df = df.copy()
        iconos_df["PDF SOURCE"] = ""
        notas = []
        procesados = []

        for file in pdf_files:
            reader = PdfReader(file)
            texto = ''.join(page.extract_text() for page in reader.pages if page.extract_text())
            if not texto.strip():
                continue

            nombre_pdf = extraer_nombre_bono_en(texto)
            if not nombre_pdf:
                continue

            nombre_pdf = nombre_pdf.upper().strip()
            if nombre_pdf not in df[columnas[0]].values:
                continue

            idx = df[df[columnas[0]] == nombre_pdf].index[0]
            fila = df.loc[idx]
            datos = extraer_datos_bono_en(texto)
            if not datos:
                continue

            errores = []
            for campo, extraido, esperado in zip(columnas[1:], datos, [fila[col] for col in columnas[1:]]):
                if not comparar_valores_bono_en(extraido, esperado):
                    errores.append(campo)
                    comentarios[(idx, campo)] = f"{campo}: In CSV: {esperado}// In PDF: {extraido}"
                    iconos_df.at[idx, campo] = f"❌ {fila[campo]}"
                else:
                    iconos_df.at[idx, campo] = f"✅ {fila[campo]}"

            if errores:
                errores_por_fila[idx] = errores
            else:
                for campo in columnas[1:]:
                    iconos_df.at[idx, campo] = f"✅ {fila[campo]}"

            iconos_df.at[idx, "PDF SOURCE"] = file.name
            procesados.append(idx)

        for idx in iconos_df.index:
            fila_notas = [comentarios[(idx, col)] for col in iconos_df.columns if (idx, col) in comentarios]
            notas.append(" | ".join(fila_notas))
        iconos_df["NOTES"] = notas

        def resaltar(row):
            idx = row.name
            return ['background-color: #FFCCCC' if col in errores_por_fila.get(idx, []) else ''
                    for col in iconos_df.columns]

        iconos_filtrados = iconos_df.loc[procesados] if procesados else pd.DataFrame()
        st.subheader("📊 Comparison Results")

        if not iconos_filtrados.empty:
            st.dataframe(iconos_filtrados.style.apply(resaltar, axis=1), use_container_width=True)
        else:
            st.warning("⚠️ No valid matches found between PDFs and CSV names.")

        # 💾 Botón para exportar Excel con errores marcados
        if st.button("📥 Create Excel", key="descargar_bono_ING"):
            if procesados:
                export_df = df.loc[procesados].copy()
                export_df["PDF SOURCE"] = iconos_df.loc[procesados]["PDF SOURCE"]
                export_df["NOTES"] = iconos_df.loc[procesados]["NOTES"]
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    ruta_excel = tmp.name
                    export_df.to_excel(ruta_excel, index=False)
                    wb = load_workbook(ruta_excel)
                    ws = wb.active
                    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    for idx in errores_por_fila:
                        excel_row = list(procesados).index(idx) + 2  # +2 por encabezado
                        for col in errores_por_fila[idx]:
                            col_idx = list(export_df.columns).index(col) + 1
                            ws.cell(row=excel_row, column=col_idx).fill = red_fill
                    wb.save(ruta_excel)
                with open(ruta_excel, "rb") as f:
                    st.download_button("📥 Download Compare_DeferredBonus.xlsx", f, file_name="Compare_DeferredBonus.xlsx")
            else:
                st.error("⚠️ No hay filas válidas para exportar.") 


    columnas_bono_en = ['NAME', 'DEFERRED BONUS', 'FINANCIAL FACTOR', 'TARGET BONUS', 'ANNUAL SALARY']
    csv_file_bono_en = st.file_uploader("📂 Upload your CSV file", type=["csv"], key="csv_bono_en")
    pdf_files_bono_en = st.file_uploader("📥 Upload your PDF files", type=["pdf"], accept_multiple_files=True, key="pdf_bono_en")
    if csv_file_bono_en and pdf_files_bono_en:
        comparar_bono_en(csv_file_bono_en, pdf_files_bono_en, columnas_bono_en)

